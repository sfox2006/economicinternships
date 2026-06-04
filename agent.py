"""
Young Economist Network Opportunity Roundup newsletter agent.

Drafts the Young Economist Network Opportunity Roundup using a Claude agent loop.
Saves the draft to Gmail with the spreadsheet attached
and emails Sam a notification — never sends.

Usage:
    python agent.py --newsletter economics-careers
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import time
import urllib.request
from datetime import date
from html import escape
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path

from anthropic import Anthropic, RateLimitError
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from configs import ALL_NEWSLETTERS, NewsletterConfig


# --- Configuration --------------------------------------------------------

SAM_EMAIL = os.environ.get("SAM_EMAIL", "samfoxanu@gmail.com")
MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
CLAUDE_CALL_DELAY_SECONDS = int(os.environ.get("CLAUDE_CALL_DELAY_SECONDS", "75"))
MAX_AGENT_STEPS = int(os.environ.get("MAX_AGENT_STEPS", "65"))
MAX_WEB_FETCHES = int(os.environ.get("MAX_WEB_FETCHES", "250"))
MAX_OUTPUT_TOKENS = int(os.environ.get("MAX_OUTPUT_TOKENS", "8000"))
MAX_COST_USD = float(os.environ.get("MAX_COST_USD", "10.00"))
INPUT_COST_PER_MTOK = float(os.environ.get("INPUT_COST_PER_MTOK", "3.00"))
OUTPUT_COST_PER_MTOK = float(os.environ.get("OUTPUT_COST_PER_MTOK", "15.00"))
MAX_FETCH_CHARS = int(os.environ.get("MAX_FETCH_CHARS", "1200"))
RATE_LIMIT_RETRY_SECONDS = int(os.environ.get("RATE_LIMIT_RETRY_SECONDS", "95"))
KEEP_RECENT_TOOL_RESULT_MESSAGES = int(os.environ.get("KEEP_RECENT_TOOL_RESULT_MESSAGES", "4"))
ROOT = Path(__file__).parent
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
TODAY = date.today().isoformat()

GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.compose",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.readonly",
]


# --- Gmail helpers --------------------------------------------------------

def gmail_service():
    creds = Credentials(
        token=None,
        refresh_token=os.environ["GMAIL_REFRESH_TOKEN"],
        client_id=os.environ["GMAIL_CLIENT_ID"],
        client_secret=os.environ["GMAIL_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=GMAIL_SCOPES,
    )
    creds.refresh(GoogleRequest())
    return build("gmail", "v1", credentials=creds, cache_discovery=False)


def gmail_search(query: str, max_results: int = 5) -> list[dict]:
    svc = gmail_service()
    res = svc.users().threads().list(userId="me", q=query, maxResults=max_results).execute()
    return res.get("threads", [])


def gmail_get_thread(thread_id: str) -> str:
    svc = gmail_service()
    thread = svc.users().threads().get(userId="me", id=thread_id, format="full").execute()
    snippets = []
    for msg in thread.get("messages", []):
        snippets.append(extract_body(msg["payload"]))
    return "\n\n---\n\n".join(snippets)


def extract_body(payload) -> str:
    if payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", "replace")
    parts = payload.get("parts", []) or []
    out = []
    for p in parts:
        if p.get("mimeType") == "text/plain":
            out.append(extract_body(p))
        elif p.get("parts"):
            out.append(extract_body(p))
    return "\n".join(out)


def create_gmail_draft(
    subject: str,
    body: str,
    attachment_path: Path | None = None,
    html_body: str | None = None,
) -> str:
    msg = MIMEMultipart("mixed")
    msg["to"] = SAM_EMAIL
    msg["from"] = SAM_EMAIL
    msg["subject"] = subject

    if html_body:
        alternative = MIMEMultipart("alternative")
        alternative.attach(MIMEText(body, "plain", "utf-8"))
        alternative.attach(MIMEText(html_body, "html", "utf-8"))
        msg.attach(alternative)
    else:
        msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachment_path and attachment_path.exists():
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(attachment_path.read_bytes())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_path.name}"')
        msg.attach(part)
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    svc = gmail_service()
    draft = svc.users().drafts().create(userId="me", body={"message": {"raw": raw}}).execute()
    return draft["id"]


def send_notification_email(drafts: list[dict]):
    """drafts is a list of {newsletter, draft_id, count} entries."""
    lines = [f"Your monthly newsletter drafts are ready for review.\n"]
    for d in drafts:
        if d.get("error"):
            lines.append(f"  ❌ {d['newsletter']}: failed — {d['error']}")
        else:
            url = f"https://mail.google.com/mail/u/0/#drafts/{d['draft_id']}"
            lines.append(f"  ✅ {d['newsletter']}: {d['count']} programs — {url}")
    lines.append(f"\nDate generated: {TODAY}")
    lines.append("Open Gmail Drafts to review, edit, and send when ready.")
    body = "\n".join(lines)

    msg = MIMEText(body, "plain", "utf-8")
    msg["to"] = SAM_EMAIL
    msg["from"] = SAM_EMAIL
    msg["subject"] = f"[Bot] Newsletter drafts ready — {TODAY}"
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    svc = gmail_service()
    svc.users().messages().send(userId="me", body={"raw": raw}).execute()


# --- Web fetch ------------------------------------------------------------

def web_fetch(url: str, max_chars: int = MAX_FETCH_CHARS) -> str:
    try:
        req = urllib.request.Request(
            url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; SamNewsletterBot/1.0)"},
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = resp.read(200_000).decode("utf-8", "replace")
        text = re.sub(r"<script[^>]*>.*?</script>", "", data, flags=re.S | re.I)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.S | re.I)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:max_chars]
    except Exception as exc:
        return f"[FETCH ERROR: {exc}]"


# --- Spreadsheet ----------------------------------------------------------

def build_spreadsheet(programs: list[dict], out_path: Path, cfg: NewsletterConfig):
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Programs"

    ws.append(cfg.spreadsheet_headers)

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(cfg.spreadsheet_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_font = Font(name="Calibri", size=11)
    for p in programs:
        ws.append([p.get(field, "") for field in cfg.spreadsheet_field_order])

    for r in range(2, len(programs) + 2):
        for c in range(1, len(cfg.spreadsheet_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    for col, w in cfg.column_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 30
    for r in range(2, len(programs) + 2):
        ws.row_dimensions[r].height = 50

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# --- Email body assembly --------------------------------------------------

def build_email_body(payload: dict, cfg: NewsletterConfig) -> str:
    intro = payload.get("intro_block", "").strip()
    closing = payload.get("closing_block", "").strip()
    programs = payload.get("programs", [])
    upcoming = payload.get("upcoming_programs", [])

    grouped: dict[str, list[dict]] = {}
    for p in programs:
        key = p.get(cfg.section_field, "Other")
        grouped.setdefault(key, []).append(p)

    sections = []
    seen = set()
    for sec_name in cfg.section_order:
        if sec_name in grouped:
            sections.append(format_section(sec_name, grouped[sec_name], cfg))
            seen.add(sec_name)
    for sec_name, items in grouped.items():
        if sec_name not in seen:
            sections.append(format_section(sec_name, items, cfg))

    if upcoming:
        sections.append(format_upcoming_section(upcoming))

    return intro + "\n\n---\n\n" + "\n\n---\n\n".join(sections) + "\n\n---\n\n" + closing


def format_section(name: str, items: list[dict], cfg: NewsletterConfig) -> str:
    emoji = cfg.section_emojis.get(name, "🌐")
    header = f"{emoji} {name.upper()}"
    blocks = [header, ""]
    for p in sorted(items, key=program_sort_key):
        blocks.append(f"{p['organisation']} — {p['program_name']} | {p['deadline']}")
        blocks.append("")
        blocks.append(p["description_paragraph"].strip())
        blocks.append("")
        blocks.append(p["url"])
        blocks.append("")
    return "\n".join(blocks).rstrip()


def program_sort_key(program: dict) -> tuple[int, str, str]:
    """Keep graduate programs/jobs at the bottom of each section."""
    type_name = program.get("type", "")
    graduate_rank = 1 if type_name in {"Graduate Program", "Graduate Job"} else 0
    return (
        graduate_rank,
        program.get("deadline", "zzzz"),
        program.get("organisation", ""),
    )


def format_upcoming_section(items: list[dict]) -> str:
    blocks = ["🗓️ COMING UP SOON", ""]
    for p in items:
        blocks.append(f"{p['organisation']} — {p['program_name']} | Expected: {p['expected_window']}")
        blocks.append("")
        blocks.append(p["description_paragraph"].strip())
        blocks.append("")
        blocks.append(p["url"])
        blocks.append("")
    return "\n".join(blocks).rstrip()


def body_to_html(body: str) -> str:
    parts = [
        '<div style="font-family: Arial, Helvetica, sans-serif; font-size: 14px; line-height: 1.5; color: #202124;">'
    ]
    for raw_line in body.splitlines():
        line = raw_line.strip()
        if not line:
            parts.append('<div style="height: 10px;"></div>')
            continue
        if line == "---":
            parts.append('<hr style="border: 0; border-top: 1px solid #dadce0; margin: 22px 0;">')
            continue
        if is_section_header(line):
            parts.append(
                f'<h2 style="font-size: 21px; line-height: 1.3; font-weight: 700; '
                f'margin: 24px 0 10px; color: #111827;">{escape(line)}</h2>'
            )
            continue
        if " — " in line and " | " in line:
            parts.append(
                f'<p style="font-size: 15px; font-weight: 700; margin: 14px 0 4px;">'
                f'{escape(line)}</p>'
            )
            continue
        if line.startswith("http://") or line.startswith("https://"):
            safe_url = escape(line, quote=True)
            parts.append(f'<p style="margin: 4px 0 12px;"><a href="{safe_url}">{escape(line)}</a></p>')
            continue
        parts.append(f'<p style="margin: 0 0 8px;">{escape(line)}</p>')
    parts.append("</div>")
    return "\n".join(parts)


def is_section_header(line: str) -> bool:
    if line == "🗓️ COMING UP SOON":
        return True
    text = re.sub(r"^[^\w]+", "", line).strip()
    return bool(text) and text == text.upper() and any(ch.isalpha() for ch in text)


# --- Agent loop -----------------------------------------------------------

def load_skill(cfg: NewsletterConfig) -> str:
    skill_dir = ROOT / cfg.skill_dir
    parts = []
    for path in sorted(skill_dir.rglob("*.md")):
        rel = path.relative_to(skill_dir)
        if rel.as_posix() == "references/organisations.md":
            parts.append(
                "=== references/organisations.md ===\n\n"
                "The full employer/program source list is available through "
                "the read_reference_file tool. Read it before checking live pages."
            )
            continue
        parts.append(f"=== {rel} ===\n\n{path.read_text(encoding='utf-8')}")
    return "\n\n".join(parts)


def read_reference_file(cfg: NewsletterConfig, filename: str) -> str:
    safe_name = Path(filename).name
    path = ROOT / cfg.skill_dir / "references" / safe_name
    allowed = {
        "organisations.md",
        "known-traps.md",
        "spreadsheet-schema.md",
        "template.md",
    }
    if safe_name not in allowed or not path.exists():
        return f"[UNKNOWN REFERENCE FILE: {filename}]"
    return path.read_text(encoding="utf-8")[:24000]


SYSTEM_PROMPT_TEMPLATE = """You are an assistant that drafts Sam Fox's "{subject}" newsletter.

Below are the skill files that define exactly how to do this. Follow them carefully.

{skill_content}

Today's date is {today}.

YOUR JOB:

1. Call gmail_search with query `subject:"{subject_root}" OR subject:"Economics Careers"` to find what's been featured recently, including editions sent under the old newsletter name.
2. Optionally call gmail_get_thread on the most recent thread to read what was in the last newsletter.
3. Call read_reference_file for `organisations.md`, then run an extensive check across the remaining organisation list. Cover every major sector and prioritise named organisations with direct careers, student, graduate, internship, cadetship, or vacationer pages.
4. Prioritise official early-career pages with direct evidence of open internships, graduate jobs, cadetships, vacationer programs, industry placements, or scholarships. Use the prior Gmail newsletter, seasonal timing, and prominent employers to choose candidates, but do not stop after only a small sample.
5. Also include a short `upcoming_programs` watchlist for opportunities likely to open in the next 3 months. These must be clearly labelled as upcoming/expected and must not be mixed into the currently-open `programs` list.
6. Submit only after you have made a broad pass through the likely sources, exhausted the useful official pages, or reached the available web-fetch/agent-step budget. Stay within the configured cost budget even if that means a shorter but still useful issue.
7. When you have a verified list of currently-open programs and any useful upcoming watchlist items, call submit_newsletter with the final structured payload.

CRITICAL RULES:
- Closed programs are EXCLUDED — not in the email body, not in the spreadsheet.
- When uncertain whether a program is open, OMIT IT. Better a shorter accurate newsletter than a long one with broken listings.
- Do not remove or suppress a currently-open program just because it appeared in any previous newsletter. Repeat it for however many newsletters are needed while the deadline is still ahead, applications are still open, or it remains useful to readers.
- Upcoming programs are allowed only in `upcoming_programs`, not `programs`. Include them when the live page or a reliable prior-cycle pattern suggests applications are likely to open within the next 3 months.
- Never invent a deadline. Use the hedging language from template.md.
- Do not use the word "genuinely".
- Match the section order, type order, header format, and tone from template.md.
- Use web_fetch efficiently. Request multiple tool calls in one assistant turn where possible, and use the available web-fetch budget for breadth before finalising.

Once you call submit_newsletter you are done.
"""


def build_submit_tool_schema(cfg: NewsletterConfig) -> dict:
    """Build the submit_newsletter tool schema dynamically per newsletter."""
    program_props = {
        "organisation": {"type": "string"},
        "program_name": {"type": "string"},
        "type": {
            "type": "string",
            "enum": [
                "Graduate Program",
                "Graduate Job",
                "Cadetship",
                "Vacationer Program",
                "Summer Vacation",
                "Internship",
                "Industry Placement",
                "Scholarship",
                "Other",
            ],
        },
        "deadline": {"type": "string"},
        "duration": {"type": "string"},
        "paid": {"type": "string"},
        "notes": {"type": "string"},
        "url": {"type": "string"},
        "description_paragraph": {"type": "string"},
    }
    required = ["organisation", "program_name", "type", "deadline",
                "duration", "paid", "url", "description_paragraph"]
    upcoming_props = {
        "organisation": {"type": "string"},
        "program_name": {"type": "string"},
        "sector": {"type": "string", "enum": cfg.section_order + ["Other"]},
        "expected_window": {"type": "string"},
        "description_paragraph": {"type": "string"},
        "url": {"type": "string"},
    }

    if cfg.section_field == "country":
        program_props["country"] = {"type": "string"}
        program_props["international"] = {
            "type": "string",
            "enum": ["Yes", "No", "Some restrictions"],
        }
        required += ["country", "international"]
    else:
        program_props["sector"] = {"type": "string", "enum": cfg.section_order + ["Other"]}
        program_props["location"] = {"type": "string"}
        program_props["citizenship"] = {
            "type": "string",
            "enum": ["Yes", "No", "Some restrictions"],
        }
        program_props["year_of_study"] = {"type": "string"}
        required += ["sector", "location", "citizenship", "year_of_study"]

    return {
        "name": "submit_newsletter",
        "description": "Submit the final newsletter payload. Call this exactly once when done.",
        "input_schema": {
            "type": "object",
            "properties": {
                "intro_block": {"type": "string"},
                "programs": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": program_props,
                        "required": required,
                    },
                },
                "upcoming_programs": {
                    "type": "array",
                    "description": (
                        "Optional watchlist of programs likely to open in the next 3 months. "
                        "Do not include these in the spreadsheet or main currently-open sections."
                    ),
                    "items": {
                        "type": "object",
                        "properties": upcoming_props,
                        "required": [
                            "organisation",
                            "program_name",
                            "sector",
                            "expected_window",
                            "description_paragraph",
                            "url",
                        ],
                    },
                },
                "closing_block": {"type": "string"},
            },
            "required": ["intro_block", "programs", "closing_block"],
        },
    }


COMMON_TOOLS = [
    {
        "name": "read_reference_file",
        "description": "Read one newsletter reference file, especially organisations.md.",
        "input_schema": {
            "type": "object",
            "properties": {"filename": {"type": "string"}},
            "required": ["filename"],
        },
    },
    {
        "name": "gmail_search",
        "description": "Search Sam's Gmail. Returns a list of matching threads.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string"},
                "max_results": {"type": "integer", "default": 5},
            },
            "required": ["query"],
        },
    },
    {
        "name": "gmail_get_thread",
        "description": "Read full plaintext body of a Gmail thread.",
        "input_schema": {
            "type": "object",
            "properties": {"thread_id": {"type": "string"}},
            "required": ["thread_id"],
        },
    },
    {
        "name": "web_fetch",
        "description": "Fetch the visible text of a web page. Use for verifying live application pages.",
        "input_schema": {
            "type": "object",
            "properties": {"url": {"type": "string"}},
            "required": ["url"],
        },
    },
]


def run_tool(name: str, args: dict) -> str:
    try:
        if name == "read_reference_file":
            return "[CONFIG ERROR: read_reference_file needs newsletter context]"
        if name == "gmail_search":
            results = gmail_search(args["query"], args.get("max_results", 5))
            return json.dumps([{"id": t["id"], "snippet": t.get("snippet", "")} for t in results])[:4000]
        if name == "gmail_get_thread":
            return gmail_get_thread(args["thread_id"])[:5000]
        if name == "web_fetch":
            return web_fetch(args["url"])
        return f"[Unknown tool: {name}]"
    except Exception as exc:
        return f"[TOOL ERROR: {exc}]"


def response_cost_usd(resp) -> float:
    usage = getattr(resp, "usage", None)
    if usage is None:
        return 0.0

    input_tokens = getattr(usage, "input_tokens", 0) or 0
    output_tokens = getattr(usage, "output_tokens", 0) or 0
    cache_creation = getattr(usage, "cache_creation_input_tokens", 0) or 0
    cache_read = getattr(usage, "cache_read_input_tokens", 0) or 0

    input_cost = (input_tokens + cache_creation) * INPUT_COST_PER_MTOK / 1_000_000
    cache_read_cost = cache_read * INPUT_COST_PER_MTOK * 0.1 / 1_000_000
    output_cost = output_tokens * OUTPUT_COST_PER_MTOK / 1_000_000
    return input_cost + cache_read_cost + output_cost


def compact_old_tool_results(messages: list[dict]) -> None:
    tool_result_message_indices = [
        i for i, msg in enumerate(messages)
        if msg.get("role") == "user"
        and isinstance(msg.get("content"), list)
        and any(
            isinstance(item, dict) and item.get("type") == "tool_result"
            for item in msg["content"]
        )
    ]
    old_indices = tool_result_message_indices[:-KEEP_RECENT_TOOL_RESULT_MESSAGES]
    for idx in old_indices:
        for item in messages[idx]["content"]:
            if not isinstance(item, dict) or item.get("type") != "tool_result":
                continue
            content = item.get("content")
            if isinstance(content, str) and len(content) > 500:
                item["content"] = content[:500] + "\n\n[Older tool result truncated to stay under rate limits.]"


def run_agent(cfg: NewsletterConfig) -> dict:
    client = Anthropic()
    subject_root = cfg.subject.rstrip("!")
    system = SYSTEM_PROMPT_TEMPLATE.format(
        subject=cfg.subject,
        subject_root=subject_root,
        skill_content=load_skill(cfg),
        today=TODAY,
    )
    tools = COMMON_TOOLS + [build_submit_tool_schema(cfg)]
    messages = [{"role": "user", "content": f"Please draft this month's {cfg.subject} newsletter."}]
    web_fetches = 0
    estimated_cost = 0.0
    budget_warning_sent = False

    for step in range(MAX_AGENT_STEPS):
        if estimated_cost >= MAX_COST_USD and budget_warning_sent:
            raise RuntimeError(
                f"Cost budget reached before a draft was submitted: "
                f"${estimated_cost:.2f} USD >= ${MAX_COST_USD:.2f} USD."
            )
        if estimated_cost >= MAX_COST_USD * 0.85 and not budget_warning_sent:
            messages.append({
                "role": "user",
                "content": (
                    f"You are near the run cost budget (${estimated_cost:.2f} USD used "
                    f"of ${MAX_COST_USD:.2f} USD). Stop researching now. Submit the "
                    "best newsletter using only programs you have already verified as "
                    "currently open. Do not call web_fetch again unless absolutely "
                    "required to complete an already verified listing."
                ),
            })
            budget_warning_sent = True

        if step:
            print(f"[{TODAY}] Waiting {CLAUDE_CALL_DELAY_SECONDS}s to stay under API rate limits...")
            time.sleep(CLAUDE_CALL_DELAY_SECONDS)
        print(f"[{TODAY}] Agent step {step + 1}/{MAX_AGENT_STEPS}...")
        compact_old_tool_results(messages)
        for attempt in range(3):
            try:
                resp = client.messages.create(
                    model=MODEL,
                    max_tokens=MAX_OUTPUT_TOKENS,
                    system=system,
                    tools=tools,
                    messages=messages,
                )
                break
            except RateLimitError:
                if attempt == 2:
                    raise
                wait = RATE_LIMIT_RETRY_SECONDS * (attempt + 1)
                print(f"[{TODAY}] Anthropic rate limit hit. Waiting {wait}s before retry...")
                time.sleep(wait)
        estimated_cost += response_cost_usd(resp)
        print(f"[{TODAY}] Estimated Claude spend: ${estimated_cost:.2f} USD / ${MAX_COST_USD:.2f} USD")
        messages.append({"role": "assistant", "content": resp.content})

        if resp.stop_reason == "max_tokens":
            messages.append({
                "role": "user",
                "content": (
                    "Your previous response hit the output token limit before completing. "
                    "Continue concisely from where you left off. If you have enough verified "
                    "open programs, call submit_newsletter now instead of doing more research."
                ),
            })
            continue

        if resp.stop_reason != "tool_use":
            raise RuntimeError(f"Agent stopped without submitting. stop_reason={resp.stop_reason}")

        tool_results = []
        submitted = None
        for block in resp.content:
            if block.type == "tool_use":
                if block.name == "submit_newsletter":
                    submitted = block.input
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": "Newsletter submitted. Thank you.",
                    })
                else:
                    if block.name == "read_reference_file":
                        result = read_reference_file(cfg, block.input["filename"])
                    elif block.name == "web_fetch":
                        web_fetches += 1
                        if web_fetches > MAX_WEB_FETCHES:
                            result = (
                                "[WEB FETCH BUDGET EXHAUSTED] Submit the best verified "
                                "newsletter now using only already-verified open programs."
                            )
                        else:
                            result = run_tool(block.name, block.input)
                    else:
                        result = run_tool(block.name, block.input)
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result,
                    })

        messages.append({"role": "user", "content": tool_results})
        if submitted:
            return submitted

    raise RuntimeError("Agent did not submit a newsletter within step budget.")


# --- Per-newsletter driver ------------------------------------------------

def run_newsletter(cfg: NewsletterConfig) -> dict:
    print(f"[{TODAY}] Starting: {cfg.name}")
    payload = run_agent(cfg)
    n = len(payload.get("programs", []))
    print(f"[{TODAY}] {cfg.name}: agent returned {n} programs.")

    sheet_path = OUTPUT_DIR / f"{cfg.filename_prefix}-{TODAY}.xlsx"
    build_spreadsheet(payload["programs"], sheet_path, cfg)
    print(f"[{TODAY}] {cfg.name}: spreadsheet saved -> {sheet_path}")

    body = build_email_body(payload, cfg)
    draft_id = create_gmail_draft(cfg.subject, body, sheet_path, body_to_html(body))
    print(f"[{TODAY}] {cfg.name}: Gmail draft created -> {draft_id}")
    return {"newsletter": cfg.name, "draft_id": draft_id, "count": n}


# --- Main -----------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--newsletter",
        choices=list(ALL_NEWSLETTERS.keys()) + ["all"],
        default="all",
        help="Which newsletter to draft. 'all' drafts both.",
    )
    args = parser.parse_args()

    targets = (list(ALL_NEWSLETTERS.values())
               if args.newsletter == "all"
               else [ALL_NEWSLETTERS[args.newsletter]])

    results = []
    for cfg in targets:
        try:
            results.append(run_newsletter(cfg))
        except Exception as exc:
            print(f"[{TODAY}] {cfg.name}: FAILED — {exc}")
            results.append({"newsletter": cfg.name, "error": str(exc)})

    send_notification_email(results)
    print(f"[{TODAY}] Notification email sent to {SAM_EMAIL}.")
    if any(result.get("error") for result in results):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
